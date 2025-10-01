#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import Set

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def normalize(s: str, case_insensitive: bool) -> str:
    s = s.strip()
    return s.lower() if case_insensitive else s


def read_targets(lines_path: Path, case_insensitive: bool) -> Set[str]:
    targets = set()
    with lines_path.open("r", encoding="utf-8") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            targets.add(normalize(s, case_insensitive))
    return targets


def highlight_rows(
    xlsx_path: Path,
    out_path: Path,
    sheet_name: str | None,
    lines_path: Path,
    color_hex: str,
    case_insensitive: bool,
    has_header: bool,
) -> None:
    targets = read_targets(lines_path, case_insensitive)
    if not targets:
        print("[WARN] No targets found in the lines file (after stripping). Nothing to do.")
        return

    wb = load_workbook(filename=str(xlsx_path))
    ws = wb[sheet_name] if sheet_name else wb.active

    # Prepare fill (solid)
    hex_clean = color_hex.replace("#", "").upper()
    if len(hex_clean) == 6:
        hex_clean = "FF" + hex_clean  # add full alpha
    fill = PatternFill(fill_type="solid", fgColor=hex_clean)

    start_row = 2 if has_header else 1
    max_row = ws.max_row
    max_col = ws.max_column

    matches = 0
    missing = 0

    # Build a quick pass over Column A to avoid re-reading cells for each target
    # Iterate rows once and mark those that match a target
    for r in range(start_row, max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is None:
            continue
        norm = normalize(str(a_val), case_insensitive)
        if norm in targets:
            # Highlight whole row (up to last used column)
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = fill
            matches += 1

    # Save
    wb.save(str(out_path))
    print(f"[INFO] Processed sheet '{ws.title}' in {xlsx_path.name}")
    print(f"[INFO] Targets loaded: {len(targets)}")
    print(f"[INFO] Rows highlighted: {matches}")
    print(f"[INFO] Output saved to: {out_path}")


def main():
    p = argparse.ArgumentParser(description="Highlight rows in XLSX when Column A matches any line from a text file.")
    p.add_argument("--xlsx", required=True, help="Path to input .xlsx file.")
    p.add_argument("--sheet", default=None, help="Worksheet name (defaults to active sheet).")
    p.add_argument("--lines", required=True, help="Path to text file with one value per line.")
    p.add_argument("--out", default=None, help="Output .xlsx (default: <input>_highlighted.xlsx).")
    p.add_argument("--color", default="#FFF59D", help="Highlight color hex (default: soft yellow #FFF59D).")
    p.add_argument("--case-insensitive", action="store_true", help="Case-insensitive matching.")
    p.add_argument("--has-header", action="store_true", help="Skip row 1 when matching (treat as header).")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    lines_path = Path(args.lines)
    out_path = Path(args.out) if args.out else xlsx_path.with_name(xlsx_path.stem + "_highlighted.xlsx")

    highlight_rows(
        xlsx_path=xlsx_path,
        out_path=out_path,
        sheet_name=args.sheet,
        lines_path=lines_path,
        color_hex=args.color,
        case_insensitive=args.case_insensitive,
        has_header=args.has_header,
    )


if __name__ == "__main__":
    main()
